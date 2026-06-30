"""Excel exporter (openpyxl). Builds a multi-sheet workbook for a project."""
import io

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


def build_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """sheets: {sheet_name: dataframe}. Returns xlsx bytes."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe = (name or "Sheet")[:31]
            (df if not df.empty else pd.DataFrame({"info": ["No data"]})).to_excel(
                writer, sheet_name=safe, index=False)
            ws = writer.sheets[safe]
            # Header styling
            header_fill = PatternFill("solid", fgColor="2563EB")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            # Auto width
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(width + 3, 48)
    buf.seek(0)
    return buf.getvalue()
